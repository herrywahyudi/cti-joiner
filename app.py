import streamlit as st
import pandas as pd
import io, zipfile
from datetime import datetime
from openpyxl import load_workbook

st.set_page_config(page_title="CTI Joiner Report", page_icon="📊", layout="wide")

# ── Helpers ───────────────────────────────────────────────────────────────────
def norm(text):
    return str(text).strip().lower().replace('_',' ').replace('-',' ').replace('/',' ')

def norm_id(v):
    return str(v).lower().replace('-','').replace('_','').replace(' ','')

def find_col(df, names):
    nc = {norm(c): c for c in df.columns}
    for n in names:
        if norm(n) in nc: return nc[norm(n)]
    return None

def clean_cell(v):
    v = str(v).strip()
    return '' if v.lower() == 'nan' else v

def clean_id(v):
    v = clean_cell(str(v))
    if not v: return ''
    try:
        n = float(v)
        if n.is_integer(): return str(int(n))
    except: pass
    return v[:-2] if v.endswith('.0') else v

JOINER_E_KEYS = ['enumbercode','enumber','seafareridnumber','seafarerid','employeeid','crewid']

CTI_OFFICES = [
    'CTI Indonesia','CTI Group Bangkok','CTI Group MCSI',
    'CTI Group Myanmar','CTI Group South Africa',
    'CTI Group Vietnam','CTI Partner Kendrick',
]
STATUSES = ['New Hire','Repeater','Re Hire','Resigned']

# ── Core functions ────────────────────────────────────────────────────────────
def read_master_file(file_bytes, filename):
    dfs = []
    if filename.endswith('.zip'):
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            for name in z.namelist():
                if name.endswith('.csv'):
                    with z.open(name) as f:
                        try: dfs.append(pd.read_csv(f))
                        except: pass
                elif name.endswith('.xlsx'):
                    with z.open(name) as f:
                        try: dfs.append(pd.read_excel(f))
                        except: pass
    elif filename.endswith('.csv'):
        dfs.append(pd.read_csv(io.BytesIO(file_bytes)))
    else:
        dfs.append(pd.read_excel(io.BytesIO(file_bytes)))
    return pd.concat(dfs, ignore_index=True) if dfs else None

def build_lookup(dfs):
    lookup = {}
    for df in dfs:
        e_col   = find_col(df, ['Seafarer ID Number','Seafarer ID','E-Number Code','Crew ID'])
        cti_col = find_col(df, ['CTI Office'])
        sta_col = find_col(df, ['Employment Status'])
        if not e_col: continue
        for _, row in df.dropna(subset=[e_col]).iterrows():
            e = clean_id(row[e_col])
            if e:
                lookup[e] = {
                    'cti':    clean_cell(row[cti_col]) if cti_col else '',
                    'status': clean_cell(row[sta_col]) if sta_col else '',
                }
    return lookup

def fill_joiner(joiner_bytes, lookup, manual_overrides):
    wb = load_workbook(io.BytesIO(joiner_bytes))
    log_lines = []
    total_filled = 0
    total_missing = []

    for sh in wb.sheetnames:
        ws = wb[sh]
        if ws.max_row < 2: continue

        headers = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(1, c).value
            if val: headers[norm_id(str(val))] = c

        e_col = None
        for k in JOINER_E_KEYS:
            if k in headers:
                e_col = headers[k]
                break
        if not e_col:
            log_lines.append(f"!! Sheet '{sh}': no E-Number column, skipped")
            continue

        cti_col_idx  = headers.get(norm_id('CTI Office'))
        stat_col_idx = headers.get(norm_id('Employment Status'))

        if not cti_col_idx:
            ins = e_col + 1
            ws.insert_cols(ins)
            ws.cell(1, ins).value = 'CTI Office'
            cti_col_idx = ins
            if stat_col_idx and stat_col_idx >= ins:
                stat_col_idx += 1

        if not stat_col_idx:
            ins = cti_col_idx + 1
            ws.insert_cols(ins)
            ws.cell(1, ins).value = 'Employment Status'
            stat_col_idx = ins

        filled = 0
        missing = []

        for r in range(2, ws.max_row + 1):
            e = clean_id(ws.cell(r, e_col).value)
            if not e: continue

            if e in manual_overrides:
                ov = manual_overrides[e]
                if ov.get('cti'):    ws.cell(r, cti_col_idx).value = ov['cti']
                if ov.get('status'): ws.cell(r, stat_col_idx).value = ov['status']
                filled += 1
            elif e in lookup:
                d = lookup[e]
                if d['cti']:    ws.cell(r, cti_col_idx).value = d['cti']
                if d['status']: ws.cell(r, stat_col_idx).value = d['status']
                filled += 1
            else:
                name_col = headers.get(norm_id('name')) or headers.get(norm_id('full name'))
                name = clean_cell(ws.cell(r, name_col).value) if name_col else ''
                missing.append({'E-Number': e, 'Name': name, 'Sheet': sh})

        log_lines.append(f"Sheet '{sh}': {filled} filled, {len(missing)} missing")
        total_filled += filled
        total_missing.extend(missing)

    log_lines.append(f"\nTotal: {total_filled} filled | {len(total_missing)} missing")
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), '\n'.join(log_lines), total_missing

def split_by_office(filled_bytes):
    xls   = pd.ExcelFile(io.BytesIO(filled_bytes))
    books = {}
    tag   = datetime.now().strftime('%Y%m%d')

    for sheet in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet)
        if 'CTI Office' not in df.columns: continue
        for office, sub in df.groupby(df['CTI Office'].fillna('UNKNOWN')):
            office = str(office).strip().replace('/', '-')
            if office.upper() == 'UNKNOWN': continue
            books.setdefault(office, {})[sheet] = sub

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for office, sheets in books.items():
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                for sname, data in sheets.items():
                    data.to_excel(writer, sheet_name=sname, index=False)
            buf.seek(0)
            zf.writestr(f'{office}_{tag}.xlsx', buf.read())

    zip_buf.seek(0)
    return zip_buf.getvalue(), sorted(books.keys())

# ── UI ────────────────────────────────────────────────────────────────────────
st.title('📊 CTI Joiner Report Tool')
st.caption('CTI Indonesia — Bali')

# ── STEP 1: Upload ────────────────────────────────────────────────────────────
st.header('Step 1 — Upload Files')
c1, c2 = st.columns(2)
with c1:
    joiner_file = st.file_uploader('Joiner Report (.xlsx)', type=['xlsx','xls'], key='joiner')
with c2:
    master_files = st.file_uploader(
        'Master Files — drop Zoho export ZIP + inactive master together',
        type=['zip','xlsx','xls','csv'],
        accept_multiple_files=True, key='masters')

st.divider()

# ── STEP 2: First Run ─────────────────────────────────────────────────────────
st.header('Step 2 — First Run')
st.caption('Run to fill what can be matched from the master. Missing rows will appear below.')

if joiner_file and master_files:
    if st.button('▶ Run First Pass', type='primary', use_container_width=True, key='run1'):
        joiner_bytes = joiner_file.read()

        # Read all master files
        dfs = []
        for f in master_files:
            df = read_master_file(f.read(), f.name)
            if df is not None:
                dfs.append(df)
                st.write(f'✓ {f.name}: {len(df):,} records')

        if not dfs:
            st.error('Could not read any master files.')
        else:
            with st.spinner('Building lookup...'):
                lookup = build_lookup(dfs)
            st.write(f'✓ Lookup ready: **{len(lookup):,}** seafarers indexed')

            with st.spinner('Filling joiner report...'):
                filled_bytes, fill_log, missing = fill_joiner(joiner_bytes, lookup, {})

            st.session_state.update({
                'joiner_bytes':  joiner_bytes,
                'lookup':        lookup,
                'fill_log_1':    fill_log,
                'missing':       missing,
                'filled_bytes':  filled_bytes,
                'step':          2 if missing else 3,
            })
else:
    st.info('Upload the Joiner Report and master files above to continue.')

# Show first run results
if 'fill_log_1' in st.session_state:
    st.code(st.session_state['fill_log_1'])

st.divider()

# ── STEP 3: Fix Missing ───────────────────────────────────────────────────────
if st.session_state.get('missing'):
    missing = st.session_state['missing']
    st.header('Step 3 — Fix Missing Rows')
    st.warning(f'⚠ {len(missing)} rows not found in master. Fill in their office and status below.')

    # Show missing table
    miss_df = pd.DataFrame(missing)
    st.dataframe(miss_df, use_container_width=True, hide_index=True)
    st.caption('E-Numbers: ' + ', '.join([m['E-Number'] for m in missing]))

    st.markdown('**Fill in the missing rows:**')

    # Pre-populate override rows from missing if not already done
    if 'overrides' not in st.session_state or \
       not any(ov['e'] for ov in st.session_state.get('overrides',[])):
        st.session_state['overrides'] = [
            {'e': m['E-Number'], 'cti': CTI_OFFICES[0], 'status': STATUSES[0]}
            for m in missing
        ]

    for i, ov in enumerate(st.session_state['overrides']):
        c1, c2, c3, c4 = st.columns([2, 2, 2, 0.5])
        with c1:
            st.session_state['overrides'][i]['e'] = st.text_input(
                'E-Number', value=ov['e'], key=f'ov_e_{i}')
        with c2:
            idx = CTI_OFFICES.index(ov['cti']) if ov['cti'] in CTI_OFFICES else 0
            st.session_state['overrides'][i]['cti'] = st.selectbox(
                'CTI Office', CTI_OFFICES, index=idx, key=f'ov_cti_{i}')
        with c3:
            idx2 = STATUSES.index(ov['status']) if ov['status'] in STATUSES else 0
            st.session_state['overrides'][i]['status'] = st.selectbox(
                'Status', STATUSES, index=idx2, key=f'ov_stat_{i}')
        with c4:
            st.write(''); st.write('')
            if st.button('✕', key=f'del_{i}') and len(st.session_state['overrides']) > 1:
                st.session_state['overrides'].pop(i)
                st.rerun()

    if st.button('+ Add Row', key='add_ov'):
        st.session_state['overrides'].append(
            {'e': '', 'cti': CTI_OFFICES[0], 'status': STATUSES[0]})
        st.rerun()

    if st.button('▶ Apply Overrides + Generate Final Files',
                 type='primary', use_container_width=True, key='run2'):
        manual = {
            ov['e'].strip(): {'cti': ov['cti'], 'status': ov['status']}
            for ov in st.session_state['overrides'] if ov['e'].strip()
        }
        with st.spinner('Applying overrides...'):
            filled_bytes, fill_log, still_missing = fill_joiner(
                st.session_state['joiner_bytes'],
                st.session_state['lookup'],
                manual)

        with st.spinner('Splitting by office...'):
            split_bytes, offices = split_by_office(filled_bytes)

        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        st.session_state.update({
            'final_filled':      filled_bytes,
            'final_filled_name': f'Joiner_Filled_{ts}.xlsx',
            'final_split':       split_bytes,
            'final_split_name':  f'CTI_By_Office_{ts}.zip',
            'final_log':         fill_log,
            'final_offices':     offices,
            'still_missing':     still_missing,
        })
        st.rerun()

    st.divider()

elif st.session_state.get('step') == 3 and 'missing' in st.session_state:
    # No missing rows — auto proceed to generate
    if 'filled_bytes' in st.session_state and 'final_filled' not in st.session_state:
        with st.spinner('Splitting by office...'):
            split_bytes, offices = split_by_office(st.session_state['filled_bytes'])
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        st.session_state.update({
            'final_filled':      st.session_state['filled_bytes'],
            'final_filled_name': f'Joiner_Filled_{ts}.xlsx',
            'final_split':       split_bytes,
            'final_split_name':  f'CTI_By_Office_{ts}.zip',
            'final_offices':     offices,
            'still_missing':     [],
        })

# ── STEP 4: Download ──────────────────────────────────────────────────────────
if 'final_filled' in st.session_state:
    st.header('Step 4 — Download')
    st.code(st.session_state.get('final_log',''))

    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            '⬇ Filled Joiner Report (.xlsx)',
            data=st.session_state['final_filled'],
            file_name=st.session_state['final_filled_name'],
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True, key='dl_filled')
    with c2:
        st.download_button(
            '⬇ Split by Office (.zip)',
            data=st.session_state['final_split'],
            file_name=st.session_state['final_split_name'],
            mime='application/zip',
            use_container_width=True, key='dl_split')

    if st.session_state.get('final_offices'):
        st.info('Offices: ' + ' · '.join(st.session_state['final_offices']))

    if st.session_state.get('still_missing'):
        still = st.session_state['still_missing']
        st.warning(f'⚠ {len(still)} rows still have no office assigned.')
        st.dataframe(pd.DataFrame(still), use_container_width=True, hide_index=True)

st.divider()
st.caption('CTI Indonesia · Bali · Internal Use Only')
